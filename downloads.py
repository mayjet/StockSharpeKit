"""
downloads.py — Generate downloadable files (Excel, JSON, ZIP) as bytes.
Supports multiple funds via result.funds list.
"""

from __future__ import annotations

import io
import json
import zipfile

import openpyxl
import pandas as pd

from analysis import AnalysisResult


def build_analysis_xlsx(result: AnalysisResult, L: dict) -> bytes:
    """Build analysis_results.xlsx and return bytes."""
    wb   = openpyxl.Workbook()

    # ── Input sheet ──────────────────────────────────────────────────────
    ws_in = wb.active
    ws_in.title = L["sheet_input"]
    ws_in.append([L["col_code"], L["col_start"], L["col_end"]])
    start_str = result.start_date.replace("-", "")
    end_str   = result.end_date.replace("-", "")

    for fi, fr in enumerate(result.funds):
        # Add fund name as a header row
        if len(result.funds) > 1:
            ws_in.append([f"--- {fr.fund_name} ---", "", ""])
        for i, row in fr.confirmed_df.iterrows():
            if fi == 0 and i == 0:
                ws_in.append([row.get("input", row["symbol"]), start_str, end_str])
            else:
                ws_in.append([row.get("input", row["symbol"]), "", ""])

    # ── Result sheets (one per fund) ─────────────────────────────────────
    for fi, fr in enumerate(result.funds):
        sheet_title = L["sheet_result"]
        if len(result.funds) > 1:
            sheet_title = f"{fr.fund_name}"
        # Ensure unique sheet name
        if fi == 0:
            ws_res = wb.create_sheet(sheet_title)
        else:
            ws_res = wb.create_sheet(sheet_title)

        ws_res.append([L["excel_title"],      ""])
        ws_res.append([L["excel_calc_date"],  result.timestamp])
        ws_res.append([L["excel_opt_type"],   L["excel_opt_tangent"]])
        ws_res.append([L["excel_exp_return"], f"{fr.expected_return*100:.2f} %"])
        ws_res.append([L["excel_ann_risk"],   f"{fr.volatility*100:.2f} %"])
        ws_res.append([L["excel_sharpe"],     f"{fr.sharpe:.4f}"])
        ws_res.append([])
        ws_res.append([L["col_ticker_name"], L["col_weight"]])
        for _, row in fr.weights_df.iterrows():
            ws_res.append(list(row))

    # ── Performance comparison sheet ─────────────────────────────────────
    ws_perf = wb.create_sheet(L.get("excel_perf_title", "Performance"))
    comp_header = result.comparison_df.columns.tolist()
    ws_perf.append(comp_header)
    _comp_data_start = ws_perf.max_row + 1
    for _, row in result.comparison_df.iterrows():
        ws_perf.append(list(row))

    # セル書式: 数値列に Excel フォーマットを適用
    _PCT_COLS = {L.get(k) for k in (
        "col_ann_return", "col_ann_risk", "col_cum_return", "col_var", "col_cvar"
    )}
    _NUM_COLS = {L.get(k) for k in ("col_sharpe", "col_beta", "col_ir")}
    for ci, col_name in enumerate(comp_header, start=1):
        fmt = None
        if col_name in _PCT_COLS:
            fmt = "0.00%"
        elif col_name in _NUM_COLS:
            fmt = "0.000"
        if fmt:
            for ri in range(_comp_data_start, _comp_data_start + len(result.comparison_df)):
                ws_perf.cell(row=ri, column=ci).number_format = fmt

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_resolved_xlsx(result: AnalysisResult) -> bytes:
    """Build resolved_symbols.xlsx and return bytes.
    Accepts AnalysisResult (multi-fund) or a plain DataFrame (legacy).
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if hasattr(result, "funds"):
            for fi, fr in enumerate(result.funds):
                sheet = fr.fund_name[:31] if len(result.funds) > 1 else "resolved_symbols"
                fr.confirmed_df.to_excel(writer, sheet_name=sheet, index=False)
        else:
            # Legacy: result is a DataFrame
            result.to_excel(writer, sheet_name="resolved_symbols", index=False)
    buf.seek(0)
    return buf.read()


def build_output_json(result: AnalysisResult, L: dict) -> bytes:
    """Build output.json and return bytes."""
    funds_data = []
    for fr in result.funds:
        funds_data.append({
            "fund_name":       fr.fund_name,
            "expected_return": round(fr.expected_return * 100, 4),
            "volatility":      round(fr.volatility * 100, 4),
            "sharpe":          round(fr.sharpe, 4),
            "weights": {
                sym: round(w, 6)
                for sym, w in fr.weights.items() if w > 0
            },
            "weights_table": {
                "header": fr.weights_df.columns.tolist(),
                "data":   fr.weights_df.values.tolist(),
            },
        })

    payload = {
        "timestamp":        result.timestamp,
        "portfolio_name":   result.portfolio_name,
        "start_date":       result.start_date,
        "end_date":         result.end_date,
        "funds":            funds_data,
        # Legacy single-fund fields (first fund)
        "expected_return":  funds_data[0]["expected_return"] if funds_data else 0,
        "volatility":       funds_data[0]["volatility"] if funds_data else 0,
        "sharpe":           funds_data[0]["sharpe"] if funds_data else 0,
        "weights":          funds_data[0]["weights"] if funds_data else {},
        "results_table":    funds_data[0]["weights_table"] if funds_data else {},
        "comparison_summary": {
            "header": result.comparison_df.columns.tolist(),
            "data":   result.comparison_df.values.tolist(),
        },
    }
    return json.dumps(payload, ensure_ascii=False, indent=4).encode("utf-8")


def build_zip(
    xlsx_bytes:     bytes,
    resolved_bytes: bytes,
    json_bytes:     bytes,
    chart_bytes:    dict[str, bytes],   # filename → bytes
) -> bytes:
    """Pack all download files into a single ZIP and return bytes."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("analysis_results.xlsx",  xlsx_bytes)
        zf.writestr("resolved_symbols.xlsx",  resolved_bytes)
        zf.writestr("output.json",            json_bytes)
        for fname, data in chart_bytes.items():
            zf.writestr(fname, data)
    buf.seek(0)
    return buf.read()
